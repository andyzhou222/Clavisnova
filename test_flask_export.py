#!/usr/bin/env python3
"""
测试Flask应用中的Excel导出
"""

import sys
import os
from pathlib import Path

# 添加backend到路径
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'backend'))

def test_flask_imports():
    """测试Flask应用中的导入"""
    print("🧪 测试Flask应用导入环境...")

    try:
        # 模拟Flask应用启动时的导入
        from config import settings
        print("✅ config imported")

        from models import Registration, Requirements
        print("✅ models imported")

        # 测试openpyxl导入
        print("Testing openpyxl import...")
        try:
            from io import BytesIO
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            print("✅ openpyxl imported successfully")

            # 测试创建工作簿
            wb = Workbook()
            ws = wb.active
            ws['A1'] = 'Test'
            print("✅ Workbook creation works")

        except ImportError as e:
            print(f"❌ openpyxl import failed: {e}")
            import traceback
            traceback.print_exc()

        # 测试SQLAlchemy
        from sqlalchemy import create_engine, desc
        from sqlalchemy.orm import sessionmaker
        print("✅ SQLAlchemy imported")

        # 测试数据库连接
        db_path = Path(__file__).parent / 'backend' / 'data' / 'Clavisnova.db'
        db_url = f"sqlite:///{db_path}"
        print(f"Database URL: {db_url}")

        engine = create_engine(db_url)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        print("✅ Database engine created")

        db = SessionLocal()
        try:
            # 测试查询
            registrations = (
                db.query(Registration)
                .order_by(desc(Registration.created_at))
                .all()
            )
            print(f"✅ Found {len(registrations)} registrations")

        finally:
            db.close()

        print("🎉 All imports work correctly!")

    except Exception as e:
        print(f"❌ Test failed: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    test_flask_imports()



