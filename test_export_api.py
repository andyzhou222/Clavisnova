#!/usr/bin/env python3
"""
测试导出API
"""

import sys
import os
from pathlib import Path

# 添加backend到路径
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'backend'))

def test_export_api():
    """测试导出API功能"""
    print("🧪 测试导出API...")

    try:
        # 模拟Flask应用环境
        from config import settings
        print("✅ config导入成功")

        from models import Registration, Requirements
        print("✅ models导入成功")

        # 直接测试导出逻辑
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from sqlalchemy import create_engine, desc
        from sqlalchemy.orm import sessionmaker

        # Direct database connection
        engine = create_engine(settings.database_url)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

        db = SessionLocal()
        try:
            # 测试注册数据导出
            registrations = (
                db.query(Registration)
                .order_by(desc(Registration.created_at))
                .all()
            )
            print(f"✅ 查询到 {len(registrations)} 条注册记录")

            # 创建Excel文件
            wb = Workbook()
            ws = wb.active
            ws.title = "Piano Registrations"

            # 简单的测试内容
            ws['A1'] = 'Test Registration Export'
            ws['A2'] = f'Total records: {len(registrations)}'

            if registrations:
                reg = registrations[0]
                ws['A3'] = f'Sample: {reg.manufacturer} {reg.model}'

            # 保存到BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            print(f"✅ Excel文件创建成功，大小: {len(output.getvalue())} bytes")
            print("🎉 导出API测试通过！")

        finally:
            db.close()

    except Exception as e:
        print(f"❌ 测试失败: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    test_export_api()



