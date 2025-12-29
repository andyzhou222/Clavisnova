#!/usr/bin/env python3
"""
调试导出API
"""

import sys
import os
from pathlib import Path

# 添加backend到路径
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'backend'))

def debug_export():
    """调试导出功能"""
    print("🔍 调试导出API...")

    try:
        # 测试所有导入
        from config import settings
        print("✅ config导入成功")

        from models import Registration, Requirements, create_tables
        print("✅ models导入成功")

        from database import db_manager
        print("✅ database导入成功")

        # 测试数据库连接
        db = db_manager.get_db()
        print("✅ 数据库连接成功")

        try:
            # 测试查询
            from sqlalchemy import desc
            registrations = (
                db.query(Registration)
                .order_by(desc(Registration.created_at))
                .all()
            )
            print(f"✅ 查询到 {len(registrations)} 条注册记录")

            # 测试Excel创建
            from io import BytesIO
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = "Test Export"

            # 简单的测试
            ws['A1'] = 'Test'
            ws['B1'] = 'Data'

            output = BytesIO()
            wb.save(output)
            output.seek(0)
            print("✅ Excel文件创建成功")

            print("🎉 所有测试通过！")

        finally:
            db.close()

    except Exception as e:
        print(f"❌ 错误: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    debug_export()



