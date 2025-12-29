#!/usr/bin/env python3
"""
简单测试导出功能
"""

import sys
import os
from pathlib import Path

# 添加backend到路径
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'backend'))

def test_export():
    """测试导出功能"""
    print("🧪 测试导出功能...")

    try:
        # 导入所需模块
        from config import settings
        from models import Registration, Requirements
        from sqlalchemy import create_engine, desc
        from sqlalchemy.orm import sessionmaker
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        print("✅ 所有导入成功")

        # 创建数据库连接 - 使用正确的backend数据库
        correct_db_path = Path(__file__).parent / 'backend' / 'data' / 'Clavisnova.db'
        db_url = f"sqlite:///{correct_db_path}"
        print(f"Using database: {db_url}")
        engine = create_engine(db_url)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

        db = SessionLocal()
        try:
            # 测试查询
            registrations = (
                db.query(Registration)
                .order_by(desc(Registration.created_at))
                .all()
            )
            print(f"✅ 查询到 {len(registrations)} 条注册记录")

            # 创建Excel文件
            wb = Workbook()
            ws = wb.active
            ws.title = "Test Export"

            # 添加一些测试数据
            headers = ["ID", "Manufacturer", "Model"]
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header)

            for row_num, reg in enumerate(registrations[:5], 2):  # 只显示前5条
                ws.cell(row=row_num, column=1, value=reg.id)
                ws.cell(row=row_num, column=2, value=reg.manufacturer or "")
                ws.cell(row=row_num, column=3, value=reg.model or "")

            # 保存到BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            print(f"✅ Excel文件创建成功，大小: {len(output.getvalue())} bytes")
            print("🎉 导出功能测试通过！")

        finally:
            db.close()

    except Exception as e:
        print(f"❌ 测试失败: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    test_export()
