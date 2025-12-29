#!/usr/bin/env python3
"""
测试Excel导出功能
"""

import sys
import os
from pathlib import Path

# 添加backend到路径
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'backend'))

def test_excel_export():
    """测试Excel导出功能"""
    print("🧪 测试Excel导出功能...")

    try:
        # 测试导入
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        print("✅ openpyxl导入成功")

        # 创建测试Excel文件
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Export"

        # 添加表头
        headers = ["ID", "Name", "Value"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # 添加测试数据
        test_data = [
            [1, "Test Item 1", "Value 1"],
            [2, "Test Item 2", "Value 2"],
            [3, "Test Item 3", "Value 3"]
        ]

        for row_num, row_data in enumerate(test_data, 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        # 保存到文件
        output_path = "test_export.xlsx"
        wb.save(output_path)

        # 检查文件是否存在
        if os.path.exists(output_path):
            file_size = os.path.getsize(output_path)
            print(f"✅ Excel文件创建成功，大小: {file_size} bytes")

            # 清理测试文件
            os.remove(output_path)
            print("✅ 测试文件已清理")
        else:
            print("❌ Excel文件创建失败")

    except ImportError as e:
        print(f"❌ openpyxl导入失败: {e}")
        print("请安装: pip install openpyxl==3.1.2")
    except Exception as e:
        print(f"❌ 测试失败: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    test_excel_export()



