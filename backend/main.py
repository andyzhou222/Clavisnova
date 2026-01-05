from flask import Flask, request, jsonify, send_file, send_from_directory, abort
from flask_cors import CORS
from flask_mail import Mail, Message
from sqlalchemy import text
import time
import json
import os
import sys
from pathlib import Path
from typing import Optional

# Ensure backend modules are importable when running as a package
current_dir = Path(__file__).resolve().parent
if str(current_dir) not in sys.path:
    sys.path.insert(0, str(current_dir))

from config import settings
from database import db_manager
from supabase_client import create_registration as supabase_create_registration, create_requirements as supabase_create_requirements, create_contact as supabase_create_contact
from schemas import (
    RegistrationCreate, RegistrationResponse, RequirementsCreate, RequirementsResponse,
    PaginationParams, PaginatedResponse, StatsResponse, HealthResponse, ErrorResponse,
    ValidationError
)
from logger import logger_manager
from models import create_tables

# Email notification helper function
def send_notification_email(form_type: str, form_data: dict):
    """Send notification email when a form is submitted"""
    try:
        if not settings.notification_email:
            logger_manager.logger.warning("Notification email not configured, skipping email notification")
            return

        # Prepare email content based on form type
        if form_type == "registration":
            subject = "🎹 新的钢琴捐赠登记 - Clavisnova"
            body = f"""
亲爱的管理员，

您收到了一份新的钢琴捐赠登记：

捐赠信息：
- 制造商: {form_data.get('manufacturer', 'N/A')}
- 型号: {form_data.get('model', 'N/A')}
- 序列号: {form_data.get('serial', 'N/A')}
- 年份: {form_data.get('year', 'N/A')}
- 类型: {form_data.get('height', 'N/A')}
- 地点: {form_data.get('city_state', 'N/A')}
- 联系方式: {form_data.get('access', 'N/A')}

请及时查看管理员后台处理此捐赠请求。

此邮件由 Clavisnova 系统自动发送。
"""

        elif form_type == "requirements":
            subject = "🎹 新的学校需求提交 - Clavisnova"
            body = f"""
亲爱的管理员，

您收到了一份新的学校钢琴需求提交：

学校信息：
- 学校名称: {form_data.get('school_name', 'N/A')}
- 现有钢琴: {form_data.get('current_pianos', 'N/A')}
- 偏好类型: {form_data.get('preferred_type', 'N/A')}
- 教师姓名: {form_data.get('teacher_name', 'N/A')}

请及时查看管理员后台处理此需求。

此邮件由 Clavisnova 系统自动发送。
"""

        elif form_type == "contact":
            subject = "🎹 新的联系表单提交 - Clavisnova"
            body = f"""
亲爱的管理员，

您收到了一份新的联系表单提交：

联系信息：
- 姓名: {form_data.get('name', 'N/A')}
- 邮箱: {form_data.get('email', 'N/A')}
- 消息内容: {form_data.get('message', 'N/A')}

请及时回复用户咨询。

此邮件由 Clavisnova 系统自动发送。
"""

        else:
            return

        # Send email
        msg = Message(
            subject=subject,
            recipients=[settings.notification_email],
            body=body
        )

        mail.send(msg)
        logger_manager.logger.info(f"Notification email sent for {form_type} submission")

    except Exception as e:
        logger_manager.logger.error(f"Failed to send notification email: {e}")
        # Don't raise exception to avoid breaking the main flow

# Initialize Flask app
# Note: Static files are served by Cloudflare Pages, not by this Flask app
app = Flask(__name__)

# Configure Flask-Mail
app.config['MAIL_SERVER'] = settings.mail_server
app.config['MAIL_PORT'] = settings.mail_port
app.config['MAIL_USE_TLS'] = settings.mail_use_tls
app.config['MAIL_USE_SSL'] = settings.mail_use_ssl
app.config['MAIL_USERNAME'] = settings.mail_username
app.config['MAIL_PASSWORD'] = settings.mail_password
app.config['MAIL_DEFAULT_SENDER'] = settings.mail_default_sender

# Initialize Flask-Mail
mail = Mail(app)

print("✅ Backend-only deployment: Frontend served by Cloudflare Pages")
print("✅ Static file serving disabled in Flask app")
print(f"✅ Mail service configured: {settings.mail_server}:{settings.mail_port}")

# CORS - enable for configured frontend origins
from flask_cors import CORS as _CORS

# Ensure FRONTEND_URL from env is included in allowed origins
frontend_url = os.getenv("FRONTEND_URL")
cors_origins = list(settings.cors_origins) if hasattr(settings, "cors_origins") else []
if frontend_url and frontend_url not in cors_origins:
    cors_origins.append(frontend_url)

# Apply CORS with credentials support
_CORS(app, origins=cors_origins, supports_credentials=True)
print(f"✅ CORS enabled for origins: {cors_origins}")

# Request logging middleware
@app.before_request
def log_request_info():
    logger_manager.logger.info(f"Request: {request.method} {request.url}")

# Test DELETE endpoint
@app.route('/api/test-delete', methods=['DELETE'])
def test_delete():
    return jsonify({"success": True, "message": "DELETE method works!"})

# API root endpoint (for API calls)
@app.route('/api/', methods=['GET'])
def api_root():
    """API root endpoint - show API info"""
    return jsonify({
        "message": "🎹 Welcome to Clavisnova Backend API",
        "version": settings.version,
        "docs": "/api/health",
        "admin": "/api/admin/stats",
        "status": "running"
    })

# Favicon endpoint
@app.route('/favicon.ico', methods=['GET'])
def favicon():
    """Favicon endpoint"""
    # Return a simple 204 No Content for favicon requests
    return '', 204

# Frontend pages - Redirect to Cloudflare Pages domain
@app.route('/', methods=['GET'])
def root_redirect():
    """Redirect root to frontend domain"""
    frontend_url = os.getenv('FRONTEND_URL', 'https://your-frontend-domain.pages.dev')
    return jsonify({
        "message": "Clavisnova Backend API",
        "frontend": frontend_url,
        "docs": "/api/health"
    })

@app.route('/index.html', methods=['GET'])
@app.route('/registration.html', methods=['GET'])
@app.route('/requirements.html', methods=['GET'])
@app.route('/admin.html', methods=['GET'])
def frontend_redirect():
    """Redirect frontend routes to Cloudflare Pages"""
    frontend_url = os.getenv('FRONTEND_URL', 'https://your-frontend-domain.pages.dev')
    return jsonify({
        "message": "Frontend served by Cloudflare Pages",
        "redirect_to": frontend_url
    }), 302

# Health check endpoint
@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    try:
        # Check database connection
        db = db_manager.get_db()
        db.execute(text("SELECT 1"))  # Simple database connectivity test
        db.close()
        db_status = "healthy"
        registration_count = 0  # Simplified for now
    except Exception as e:
        db_status = f"unhealthy: {str(e)}"
        registration_count = 0

    response = HealthResponse(
        status="healthy" if db_status == "healthy" else "unhealthy",
        timestamp=time.time(),
        version=settings.version,
        database=db_status,
        registrations=registration_count,
        uptime=time.time(),  # Simplified uptime
        memory_usage=0.0  # Simplified memory usage
    )

    return jsonify(response.__dict__)

# Registration endpoints
@app.route('/api/registration', methods=['POST'])
def create_registration():
    """Create a new registration"""
    try:
        from models import Registration

        # Log incoming request data for debugging
        logger_manager.logger.info(f"Registration request received")

        data = request.get_json()
        logger_manager.logger.info(f"Request data: {data}")

        # Ensure year is an integer
        year_value = data.get('year', 2020)
        if isinstance(year_value, str):
            try:
                year_value = int(year_value)
                logger_manager.logger.info(f"Converted year string '{data.get('year')}' to int {year_value}")
            except ValueError:
                logger_manager.logger.warning(f"Invalid year value: '{data.get('year')}', using default 2020")
                year_value = 2020

        # Create registration object
        registration = RegistrationCreate(
            manufacturer=data.get('manufacturer', ''),
            model=data.get('model', ''),
            serial=data.get('serial', ''),
            year=year_value,
            height=data.get('height', ''),
            finish=data.get('finish', ''),
            color_wood=data.get('color_wood', ''),
            access=data.get('access', ''),
            city_state=data.get('city_state', ''),
            ip_address=request.remote_addr,
            user_agent=request.headers.get('User-Agent')
        )

        logger_manager.logger.info(f"Registration object created successfully: manufacturer={registration.manufacturer}, model={registration.model}")

        # Try REST fallback (Supabase) if configured
        use_rest = os.getenv("USE_SUPABASE_REST", "false").lower() in ("1", "true", "yes")
        if use_rest:
            # map to supabase fields expected by table
            sb_payload = {
                "manufacturer": registration.manufacturer,
                "model": registration.model,
                "serial": registration.serial,
                "year": registration.year,
                "height": registration.height,
                "finish": registration.finish,
                "color_wood": registration.color_wood,
                "access": registration.access,
                "city_state": registration.city_state,
                "ip_address": registration.ip_address,
                "user_agent": registration.user_agent
            }
            try:
                created = supabase_create_registration(sb_payload)
                result_id = created.get("id")
                logger_manager.logger.info(f"Registration saved via Supabase REST with ID: {result_id}")
                response = RegistrationResponse(id=result_id, message="Registration created successfully")
                return jsonify(response.__dict__), 201
            except Exception as e:
                logger_manager.logger.error(f"Supabase REST error: {e}", exc_info=True)
                return jsonify({"message": "Supabase REST error"}), 500

        # Save to database synchronously (default)
        db = db_manager.get_db()
        try:
            reg = Registration(
                manufacturer=registration.manufacturer,
                model=registration.model,
                serial=registration.serial,
                year=registration.year,
                height=registration.height,
                finish=registration.finish,
                color_wood=registration.color_wood,
                access=registration.access,
                city_state=registration.city_state,
                ip_address=registration.ip_address,
                user_agent=registration.user_agent
            )
            db.add(reg)
            db.commit()
            db.refresh(reg)
            result_id = reg.id
            logger_manager.logger.info(f"Registration saved to database with ID: {result_id}")
        except Exception as db_error:
            logger_manager.logger.error(f"Database error during registration save: {db_error}")
            raise db_error
        finally:
            db.close()

        response = RegistrationResponse(id=result_id, message="Registration created successfully")

        # Send notification email
        notification_data = {
            'manufacturer': registration.manufacturer,
            'model': registration.model,
            'serial': registration.serial,
            'year': registration.year,
            'height': registration.height,
            'city_state': registration.city_state,
            'access': registration.access
        }
        send_notification_email("registration", notification_data)

        logger_manager.logger.info(f"Registration API completed successfully")
        return jsonify(response.__dict__), 201

    except ValidationError as e:
        logger_manager.logger.warning(f"Validation error: {e}")
        return jsonify(ErrorResponse(message=str(e)).__dict__), 400
    except Exception as e:
        logger_manager.logger.error(f"Registration error: {e}", exc_info=True)
        return jsonify(ErrorResponse(message="Internal server error").__dict__), 500

# Requirements endpoints
@app.route('/api/requirements', methods=['POST'])
def create_requirements():
    """Create a new requirements submission"""
    try:
        from models import Requirements

        data = request.get_json()

        # Create requirements object
        requirements = RequirementsCreate(
            school_name=data.get('school_name') or data.get('info1'),  # 向后兼容
            current_pianos=data.get('current_pianos') or data.get('info2'),
            preferred_type=data.get('preferred_type') or data.get('info3'),
            teacher_name=data.get('teacher_name') or data.get('info4'),
            background=data.get('background') or data.get('info5'),
            commitment=data.get('commitment') or data.get('info6'),
            ip_address=request.remote_addr,
            user_agent=request.headers.get('User-Agent')
        )

        use_rest = os.getenv("USE_SUPABASE_REST", "false").lower() in ("1", "true", "yes")
        if use_rest:
            sb_payload = {
                "school_name": requirements.school_name,
                "current_pianos": requirements.current_pianos,
                "preferred_type": requirements.preferred_type,
                "teacher_name": requirements.teacher_name,
                "background": requirements.background,
                "commitment": requirements.commitment,
                "ip_address": requirements.ip_address,
                "user_agent": requirements.user_agent
            }
            try:
                created = supabase_create_requirements(sb_payload)
                result_id = created.get("id")
                return jsonify(RequirementsResponse(id=result_id, message="Requirements submitted successfully").__dict__), 201
            except Exception as e:
                logger_manager.logger.error(f"Supabase REST error for requirements: {e}", exc_info=True)
                return jsonify({"message": "Supabase REST error"}), 500

        # Save to database synchronously
        db = db_manager.get_db()
        try:
            req = Requirements(
                school_name=requirements.school_name,
                current_pianos=requirements.current_pianos,
                preferred_type=requirements.preferred_type,
                teacher_name=requirements.teacher_name,
                background=requirements.background,
                commitment=requirements.commitment,
                ip_address=requirements.ip_address,
                user_agent=requirements.user_agent
            )
            db.add(req)
            db.commit()
            db.refresh(req)
            result_id = req.id
        finally:
            db.close()

        response = RequirementsResponse(id=result_id, message="Requirements submitted successfully")

        # Send notification email
        notification_data = {
            'school_name': requirements.school_name,
            'current_pianos': requirements.current_pianos,
            'preferred_type': requirements.preferred_type,
            'teacher_name': requirements.teacher_name
        }
        send_notification_email("requirements", notification_data)

        return jsonify(response.__dict__), 201

    except ValidationError as e:
        return jsonify(ErrorResponse(message=str(e)).__dict__), 400
    except Exception as e:
        logger_manager.logger.error(f"Requirements error: {e}")
        return jsonify(ErrorResponse(message="Internal server error").__dict__), 500

# Contact endpoints
@app.route('/api/contact', methods=['POST'])
def create_contact():
    """Create a new contact message"""
    try:
        from models import Contact

        data = request.get_json() or {}
        name = data.get('name', '')
        email = data.get('email', '')
        message_text = data.get('message', '')

        if not message_text or not message_text.strip():
            return jsonify(ErrorResponse(message="Message cannot be empty").__dict__), 400

        use_rest = os.getenv("USE_SUPABASE_REST", "false").lower() in ("1", "true", "yes")
        if use_rest:
            sb_payload = {
                "name": name,
                "email": email,
                "message": message_text,
                "ip_address": request.remote_addr,
                "user_agent": request.headers.get('User-Agent')
            }
            try:
                created = supabase_create_contact(sb_payload)
                cid = created.get("id")
                return jsonify({"id": cid, "message": "Contact submitted"}), 201
            except Exception as e:
                logger_manager.logger.error(f"Supabase REST error for contact: {e}", exc_info=True)
                return jsonify({"message": "Supabase REST error"}), 500

        db = db_manager.get_db()
        try:
            contact = Contact(
                name=name,
                email=email,
                message=message_text,
                ip_address=request.remote_addr,
                user_agent=request.headers.get('User-Agent')
            )
            db.add(contact)
            db.commit()
            db.refresh(contact)
            cid = contact.id
        finally:
            db.close()

        # Send notification email
        notification_data = {
            'name': name,
            'email': email,
            'message': message_text
        }
        send_notification_email("contact", notification_data)

        return jsonify({"id": cid, "message": "Contact submitted"}), 201
    except Exception as e:
        logger_manager.logger.error(f"Contact error: {e}", exc_info=True)
        return jsonify(ErrorResponse(message="Internal server error").__dict__), 500


@app.route('/api/admin/contacts', methods=['GET'])
def get_contacts():
    """Get all contacts (admin)"""
    try:
        from models import Contact
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 25))

        db = db_manager.get_db()
        try:
            query = db.query(Contact).order_by(Contact.created_at.desc())
            total = query.count()
            offset = (page - 1) * limit
            items = query.offset(offset).limit(limit).all()
            data = [c.to_dict() for c in items]
            total_pages = (total + limit - 1) // limit
            pagination = {
                "page": page,
                "limit": limit,
                "total": total,
                "total_pages": total_pages,
                "has_next": page < total_pages,
                "has_prev": page > 1
            }
            return jsonify({"success": True, "data": data, "pagination": pagination}), 200
        finally:
            db.close()
    except Exception as e:
        logger_manager.logger.error(f"Get contacts error: {e}")
        return jsonify({"success": False, "message": "Internal server error"}), 500

# Admin endpoints
@app.route('/api/admin/registrations', methods=['GET'])
def get_registrations():
    """Get all registrations (admin)"""
    try:
        from models import Registration
        from sqlalchemy import or_, func

        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 25))
        search = request.args.get('search', '')

        db = db_manager.get_db()
        try:
            # Build query
            query = db.query(Registration)

            # Add search filter if provided
            if search:
                search_filter = f"%{search}%"
                query = query.filter(
                    or_(
                        Registration.manufacturer.ilike(search_filter),
                        Registration.model.ilike(search_filter),
                        Registration.serial.ilike(search_filter),
                        Registration.city_state.ilike(search_filter)
                    )
                )

            # Get total count
            total = query.count()

            # Apply pagination
            offset = (page - 1) * limit
            registrations = query.offset(offset).limit(limit).all()

            # Convert to dict
            data = [reg.to_dict() for reg in registrations]

            # Calculate pagination info
            total_pages = (total + limit - 1) // limit
            pagination = {
                "page": page,
                "limit": limit,
                "total": total,
                "total_pages": total_pages,
                "has_next": page < total_pages,
                "has_prev": page > 1
            }

            # Return format expected by admin.html
            return jsonify({
                "success": True,
                "data": data,
                "pagination": pagination
            }), 200

        finally:
            db.close()

    except Exception as e:
        logger_manager.logger.error(f"Get registrations error: {e}")
        return jsonify({"success": False, "message": "Internal server error"}), 500

@app.route('/api/admin/requirements', methods=['GET'])
def get_requirements():
    """Get all requirements (admin)"""
    try:
        from models import Requirements
        from sqlalchemy import or_, func

        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 25))
        search = request.args.get('search', '')

        db = db_manager.get_db()
        try:
            # Build query
            query = db.query(Requirements)

            # Add search filter if provided (search in all text fields)
            if search:
                search_filter = f"%{search}%"
                query = query.filter(
                    or_(
                        Requirements.school_name.ilike(search_filter),
                        Requirements.current_pianos.ilike(search_filter),
                        Requirements.preferred_type.ilike(search_filter),
                        Requirements.teacher_name.ilike(search_filter),
                        Requirements.background.ilike(search_filter),
                        Requirements.commitment.ilike(search_filter)
                    )
                )

            # Get total count
            total = query.count()

            # Apply pagination
            offset = (page - 1) * limit
            requirements = query.offset(offset).limit(limit).all()

            # Convert to dict
            data = [req.to_dict() for req in requirements]

            # Calculate pagination info
            total_pages = (total + limit - 1) // limit
            pagination = {
                "page": page,
                "limit": limit,
                "total": total,
                "total_pages": total_pages,
                "has_next": page < total_pages,
                "has_prev": page > 1
            }

            # Return format expected by admin.html
            return jsonify({
                "success": True,
                "data": data,
                "pagination": pagination
            }), 200

        finally:
            db.close()

    except Exception as e:
        logger_manager.logger.error(f"Get requirements error: {e}")
        return jsonify({"success": False, "message": "Internal server error"}), 500

@app.route('/api/admin/stats', methods=['GET'])
def get_stats():
    """Get system statistics (admin)"""
    try:
        from models import Registration, Requirements
        from sqlalchemy import func

        db = db_manager.get_db()
        try:
            # Get counts
            registration_count = db.query(func.count(Registration.id)).scalar()
            requirements_count = db.query(func.count(Requirements.id)).scalar()
            total_submissions = registration_count + requirements_count

            stats = {
                "registrations": registration_count,
                "requirements": requirements_count,
                "total_submissions": total_submissions
            }

            # Return format expected by admin.html
            return jsonify({
                "success": True,
                "stats": stats
            }), 200

        finally:
            db.close()

    except Exception as e:
        logger_manager.logger.error(f"Get stats error: {e}")
        return jsonify({"success": False, "message": "Internal server error"}), 500

# Delete endpoints (using GET method with action parameter to avoid HTTP method issues)
@app.route('/api/admin/delete/registration/<id>', methods=['GET', 'OPTIONS'])
def delete_registration(id):
    """Delete a specific registration (admin)"""
    try:
        from models import Registration

        # Convert id to integer
        try:
            registration_id = int(id)
        except ValueError:
            return jsonify({"success": False, "message": "Invalid ID format"}), 400

        db = db_manager.get_db()
        try:
            registration = db.query(Registration).filter(Registration.id == registration_id).first()
            if not registration:
                return jsonify({"success": False, "message": "Registration not found"}), 404

            db.delete(registration)
            db.commit()

            return jsonify({"success": True, "message": "Registration deleted successfully"}), 200

        finally:
            db.close()

    except Exception as e:
        logger_manager.logger.error(f"Delete registration error: {e}")
        return jsonify({"success": False, "message": "Internal server error"}), 500

@app.route('/api/admin/delete/requirement/<id>', methods=['GET', 'OPTIONS'])
def delete_requirement(id):
    """Delete a specific requirement (admin)"""
    try:
        from models import Requirements

        # Convert id to integer
        try:
            requirement_id = int(id)
        except ValueError:
            return jsonify({"success": False, "message": "Invalid ID format"}), 400

        db = db_manager.get_db()
        try:
            requirement = db.query(Requirements).filter(Requirements.id == requirement_id).first()
            if not requirement:
                return jsonify({"success": False, "message": "Requirement not found"}), 404

            db.delete(requirement)
            db.commit()

            return jsonify({"success": True, "message": "Requirement deleted successfully"}), 200

        finally:
            db.close()

    except Exception as e:
        logger_manager.logger.error(f"Delete requirement error: {e}")
        return jsonify({"success": False, "message": "Internal server error"}), 500

# Delete contact
@app.route('/api/admin/delete/contact/<id>', methods=['GET', 'OPTIONS'])
def delete_contact(id):
    try:
        from models import Contact
        try:
            contact_id = int(id)
        except ValueError:
            return jsonify({"success": False, "message": "Invalid ID format"}), 400

        db = db_manager.get_db()
        try:
            contact = db.query(Contact).filter(Contact.id == contact_id).first()
            if not contact:
                return jsonify({"success": False, "message": "Contact not found"}), 404

            db.delete(contact)
            db.commit()
            return jsonify({"success": True, "message": "Contact deleted successfully"}), 200
        finally:
            db.close()
    except Exception as e:
        logger_manager.logger.error(f"Delete contact error: {e}")
        return jsonify({"success": False, "message": "Internal server error"}), 500

# Admin endpoints

# Export endpoints
@app.route('/api/admin/export/registrations', methods=['GET'])
def export_registrations():
    """Export all registrations as Excel file (or CSV fallback)"""
    try:
        # Try to import Excel modules, fall back to CSV if not available
        try:
            from io import BytesIO, StringIO
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            use_excel = True
        except ImportError:
            from io import StringIO
            use_excel = False
            print("WARNING: openpyxl not available, falling back to CSV export")

        from sqlalchemy import create_engine, desc
        from sqlalchemy.orm import sessionmaker
        from models import Registration

        # Direct database connection (avoiding db_manager logger issues)
        engine = create_engine(settings.database_url)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

        db = SessionLocal()
        try:
            registrations = (
                db.query(Registration)
                .order_by(desc(Registration.created_at))
                .all()
            )

            if use_excel:
                # Create Excel workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Piano Registrations"

                # Header styling
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")

                # Headers
                headers = [
                    "ID", "Manufacturer", "Model", "Serial #", "Year", "Type", "Height",
                    "Finish", "Condition", "Color/Wood", "City/State", "Access", "IP Address", "Created At", "Updated At"
                ]

                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.font = header_font
                    cell.fill = header_fill

                # Data rows
                for row_num, reg in enumerate(registrations, 2):
                    ws.cell(row=row_num, column=1, value=reg.id)
                    ws.cell(row=row_num, column=2, value=reg.manufacturer or "")
                    ws.cell(row=row_num, column=3, value=reg.model or "")
                    ws.cell(row=row_num, column=4, value=reg.serial or "")
                    ws.cell(row=row_num, column=5, value=reg.year or "")
                    ws.cell(row=row_num, column=6, value=reg.height or "")  # Type stored in height column
                    ws.cell(row=row_num, column=7, value=reg.height or "")  # Height/Length (if available)
                    ws.cell(row=row_num, column=8, value=reg.finish or "")
                    ws.cell(row=row_num, column=9, value=reg.finish or "")  # Condition stored in finish column
                    ws.cell(row=row_num, column=10, value=reg.color_wood or "")
                    ws.cell(row=row_num, column=11, value=reg.city_state or "")
                    ws.cell(row=row_num, column=12, value=reg.access or "")
                    ws.cell(row=row_num, column=13, value=reg.ip_address or "")
                    ws.cell(row=row_num, column=14, value=str(reg.created_at) if reg.created_at else "")
                    ws.cell(row=row_num, column=15, value=str(reg.updated_at) if reg.updated_at else "")

                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Max width 50
                    ws.column_dimensions[column_letter].width = adjusted_width

                # Save to BytesIO
                output = BytesIO()
                wb.save(output)
                output.seek(0)

                return send_file(
                    output,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name='piano_registrations.xlsx'
                )
            else:
                # Fallback to CSV
                output = StringIO()
                import csv

                writer = csv.writer(output)
                # Headers
                writer.writerow([
                    "ID", "Manufacturer", "Model", "Serial #", "Year", "Type", "Height",
                    "Finish", "Condition", "Color/Wood", "City/State", "Access", "IP Address", "Created At", "Updated At"
                ])

                # Data rows
                for reg in registrations:
                    writer.writerow([
                        reg.id,
                        reg.manufacturer or "",
                        reg.model or "",
                        reg.serial or "",
                        reg.year or "",
                        reg.height or "",  # Type stored in height
                        reg.height or "",  # Height/Length if any
                        reg.finish or "",
                        reg.finish or "",  # Condition stored in finish
                        reg.color_wood or "",
                        reg.city_state or "",
                        reg.access or "",
                        reg.ip_address or "",
                        str(reg.created_at) if reg.created_at else "",
                        str(reg.updated_at) if reg.updated_at else ""
                    ])

                output.seek(0)
                return send_file(
                    output,
                    mimetype='text/csv',
                    as_attachment=True,
                    download_name='piano_registrations.csv'
                )

        finally:
            db.close()

    except Exception as e:
        print(f"Export registrations error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": f"Export failed: {str(e)}"}), 500

@app.route('/api/admin/export/requirements', methods=['GET'])
def export_requirements():
    """Export all requirements as Excel file (or CSV fallback)"""
    try:
        # Try to import Excel modules, fall back to CSV if not available
        try:
            from io import BytesIO, StringIO
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            use_excel = True
        except ImportError:
            from io import StringIO
            use_excel = False
            print("WARNING: openpyxl not available, falling back to CSV export")

        from sqlalchemy import create_engine, desc
        from sqlalchemy.orm import sessionmaker
        from models import Requirements

        # Direct database connection (avoiding db_manager logger issues)
        engine = create_engine(settings.database_url)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

        db = SessionLocal()
        try:
            requirements = (
                db.query(Requirements)
                .order_by(desc(Requirements.created_at))
                .all()
            )

            if use_excel:
                # Create Excel workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Requirements"

                # Header styling
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="28B463", end_color="28B463", fill_type="solid")

                # Headers
                headers = [
                    "ID", "School Name", "Current Pianos", "Preferred Type", "Teacher Name", "Background", "Commitment",
                    "IP Address", "Created At", "Updated At"
                ]

                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.font = header_font
                    cell.fill = header_fill

                # Data rows
                for row_num, req in enumerate(requirements, 2):
                    ws.cell(row=row_num, column=1, value=req.id)
                    ws.cell(row=row_num, column=2, value=req.school_name or "")
                    ws.cell(row=row_num, column=3, value=req.current_pianos or "")
                    ws.cell(row=row_num, column=4, value=req.preferred_type or "")
                    ws.cell(row=row_num, column=5, value=req.teacher_name or "")
                    ws.cell(row=row_num, column=6, value=req.background or "")
                    ws.cell(row=row_num, column=7, value=req.commitment or "")
                    ws.cell(row=row_num, column=8, value=req.ip_address or "")
                    ws.cell(row=row_num, column=9, value=str(req.created_at) if req.created_at else "")
                    ws.cell(row=row_num, column=10, value=str(req.updated_at) if req.updated_at else "")

                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Max width 50
                    ws.column_dimensions[column_letter].width = adjusted_width

                # Save to BytesIO
                output = BytesIO()
                wb.save(output)
                output.seek(0)

                return send_file(
                    output,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name='requirements.xlsx'
                )
            else:
                # Fallback to CSV
                output = StringIO()
                import csv

                writer = csv.writer(output)
                # Headers
                writer.writerow([
                    "ID", "School Name", "Current Pianos", "Preferred Type", "Teacher Name", "Background", "Commitment",
                    "IP Address", "Created At", "Updated At"
                ])

                # Data rows
                for req in requirements:
                    writer.writerow([
                        req.id,
                        req.school_name or "",
                        req.current_pianos or "",
                        req.preferred_type or "",
                        req.teacher_name or "",
                        req.background or "",
                        req.commitment or "",
                        req.ip_address or "",
                        str(req.created_at) if req.created_at else "",
                        str(req.updated_at) if req.updated_at else ""
                    ])

                output.seek(0)
                return send_file(
                    output,
                    mimetype='text/csv',
                    as_attachment=True,
                    download_name='requirements.csv'
                )

        finally:
            db.close()

    except Exception as e:
        print(f"Export requirements error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": f"Export failed: {str(e)}"}), 500

# Startup and shutdown events
def startup_event():
    """Application startup tasks"""
    logger_manager.logger.info("🎹 Clavisnova Flask Backend Server Starting...")

    # Create database tables if they don't exist
    try:
        from models import create_tables
        create_tables()
        logger_manager.logger.info("✅ Database tables created/verified")
    except Exception as e:
        logger_manager.logger.error(f"❌ Failed to create database tables: {e}")
        raise e

    logger_manager.logger.info(f"📊 Version: {settings.version}")
    logger_manager.logger.info(f"🌐 Host: {settings.host}:{settings.port}")
    logger_manager.logger.info(f"📁 Database: {settings.database_url}")

def cleanup_app_context():
    """Application context cleanup tasks"""
    logger_manager.logger.info("🧹 Cleaning up Flask application context")

# Register startup/shutdown functions
with app.app_context():
    startup_event()

@app.teardown_appcontext
def teardown_appcontext(exception=None):
    cleanup_app_context()

if __name__ == '__main__':
    app.run(
        host=settings.host,
        port=settings.port,
        debug=settings.debug
    )
